import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tkinter import Tk, filedialog, messagebox
import shutil

def choose_file():
    try:
        root = Tk()
        root.withdraw()

        file_path = filedialog.askopenfilename(
            title="Выберите файл",
            filetypes=[("Excel files", "*.xlsx; *.xls")]
        )
        
        return file_path
    except Exception as e:
        print(f"Ошибка при выборе файла: {str(e)}")
        return None
    

def choose_output_directory():
    try:
        root = Tk()
        root.withdraw()

        output_directory = filedialog.askdirectory(
            title="Выберите каталог для сохранения файла с результатами"
        )

        return output_directory
    except Exception as e:
        print(f"Ошибка при выборе каталога: {str(e)}")
        return None
    

def compare_and_highlight(file1, file2):
    try:
        output_file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx; *.xls")]
        )
        shutil.copy(file1, output_file_path)

        df1 = pd.read_excel(output_file_path, header=None)
        df2 = pd.read_excel(file2, header=None)

        df1, df2 = df1.align(df2, join='outer', axis=1)

        workbook = load_workbook(output_file_path)
        sheet = workbook.active

        for idx, (row1, row2) in enumerate(zip(df1.iterrows(), df2.iterrows()), start=1):
            for col, (value1, value2) in enumerate(zip(row1[1], row2[1])):
                if pd.notna(value1) and pd.notna(value2) and value1 != value2:
                    print(f"Несоответствие в ячейках {idx}, {col + 1}: {value1} != {value2}")
                    sheet.cell(row=idx, column=col + 1).fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

        workbook.save(output_file_path)
        workbook.close()
        print("=" * 50)
        print(f"Файл с несоответствиями сохранен {output_file_path}")
        print("=" * 50)
        messagebox.showinfo("Успешно", f"Файл с несоответствиями сохранен:\n{output_file_path}")
    except Exception as e:
        print(f"Ошибка при сравнении и выделении: {str(e)}")


file1_path = choose_file()
file2_path = choose_file()

if file1_path and file2_path:
    compare_and_highlight(file1_path, file2_path)
else:
    print("Выбор файлов отменен.")